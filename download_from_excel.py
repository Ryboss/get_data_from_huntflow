import openpyxl
import os
import codecs
import httpx


def get_files_data(directory: str, name: str, token: str):
    """
    Получение данных с файла
    """
    walk = os.walk(directory)

    for i in walk:
        for file in i[2]:
            if file.split(".")[0] == name:

                if file.split(".")[-1] == "pdf":
                    with codecs.open(f"{directory}/{file}", "rb") as file_pdf:
                        request = httpx.post(
                            "https://dev-100-api.huntflow.dev/v2/accounts/17/upload",
                            files={"file": file_pdf.read()},
                            headers={
                                "Authorization": f"Bearer {token}",
                                "Content-Type": "multipart/form-data; boundary=----WebKitFormBoundaryj4XmM7UjQ6ySmxDC",
                                "x-file-parse": "true",
                                "Accept": "application/json"})

                        if request.status_code != 200:
                            raise Exception("Резюме не получено")

                else:
                    with open(f"{directory}/{file}", "rb") as file_doc:
                        request = httpx.post(
                            "https://dev-100-api.huntflow.dev/v2/accounts/17/upload",
                            files={"file": file_doc.read()},
                            headers={
                                "Authorization": f"Bearer {token}",
                                "Content-Type": "multipart/form-data; boundary=----WebKitFormBoundaryj4XmM7UjQ6ySmxDC",
                                "x-file-parse": "true",
                                "Accept": "application/json"})
                        if request.status_code != 200:
                            raise Exception("Резюме не получено")

    return request.json()


def get_data_from_excel(file_path: str, token: str) -> list:
    """
    Получение данных из Excel
    """

    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)

    # Получаем активный лист
    sheet = workbook.active

    # Создаем словарь для хранения данных по каждому пользователю
    data = {}
    candidate_list = []
    # Проходимся по каждой строке в листе
    for row in sheet.iter_rows(min_row=2):  # Начинаем со 2-й строки, чтобы пропустить заголовки
        # Заполняем поля по каждому кандидату

        data["full_name"] = " ".join(row[1].value.split())
        data["position"] = row[0].value
        data["money"] = row[2].value
        data["comments"] = row[3].value
        data["status"] = row[4].value

        full_data = get_files_data(directory=data['position'], name=data["full_name"], token=token)

        data["email"] = full_data.get('fields').get("email", None)
        data["telegram"] = full_data.get('fields').get("telegram", None)
        data["phones"] = full_data.get('fields').get("phones", None)
        data["text"] = full_data.get("text", None)
        data["files"] = [full_data.get("id")]

        if full_data.get("photo"):
            data["photo"] = full_data.get("photo").get("id")

        if full_data.get('fields').get('birthdate'):
            data["birthdate"] = f"{full_data.get('fields').get('birthdate').get('year')}-{full_data.get('fields').get('birthdate').get('month')}-{full_data.get('fields').get('birthdate').get('day')}"
        data["skype"] = full_data.get("skype", None)

        candidate_list.append(data)
        data = {}

    workbook.close()

    # Выводим информацию по каждому пользователю
    return candidate_list


def delete_rows_in_excel(file_path: str):
    """
    Удаление кандидата из базовой таблицы
    после заноса в базу и на вакансию (В случае разрыва соединения или внештатного отключения программы,
    можно будет продолжить со следующего кандидата)
    """

    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)

    # Получаем активный лист
    sheet = workbook.active

    # Удалаяем по 1 строке
    sheet.delete_rows(2, 1)

    # Сохраняем
    workbook.save(file_path)

    workbook.close()
    print("Кандидат удален из табоицы и добавлен в базу")